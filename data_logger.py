"""Moringa six-band sensor gateway, LPQI analyzer, and Excel logger.

Run this file, then open http://localhost:5050/dashboard.html.  The browser never
contacts the ESP8266 directly: this service requests a scan, validates the six
bands, computes LPQI, stores the complete record, and returns that same record
to the dashboard.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import threading
from dataclasses import dataclass
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Mapping
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request, urlopen

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


BANDS = ("R", "S", "T", "U", "V", "W")
HEADERS = (
    "Sample No.",
    "Date",
    "Time",
    "R - 610 nm",
    "S - 680 nm",
    "T - 730 nm",
    "U - 760 nm",
    "V - 810 nm",
    "W - 860 nm",
    "Red-edge avg",
    "NIR avg",
    "LPQI",
    "Quality",
    "Sensor URL",
)


class SensorError(RuntimeError):
    """A readable sensor acquisition/validation failure."""


def parse_sensor_payload(payload: str | Mapping[str, Any]) -> dict[str, float]:
    """Parse JSON or text such as ``R,95.2 S,68.5 ...`` into six bands."""
    candidate: Mapping[str, Any] | None = None
    raw_text = ""

    if isinstance(payload, Mapping):
        candidate = payload.get("data", payload)  # type: ignore[arg-type]
    else:
        raw_text = payload.strip()
        try:
            decoded = json.loads(raw_text)
            if isinstance(decoded, Mapping):
                candidate = decoded.get("data", decoded)  # type: ignore[arg-type]
        except json.JSONDecodeError:
            candidate = None

    parsed: dict[str, float] = {}
    if isinstance(candidate, Mapping):
        upper = {str(key).upper(): value for key, value in candidate.items()}
        for band in BANDS:
            if band in upper:
                try:
                    parsed[band] = float(upper[band])
                except (TypeError, ValueError):
                    pass

    if not parsed and raw_text:
        pattern = r"([RSTUVWrstuvw])\s*[,=:]\s*(-?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)"
        for band, value in re.findall(pattern, raw_text):
            parsed[band.upper()] = float(value)

    missing = [band for band in BANDS if band not in parsed]
    if missing:
        raise SensorError(f"Incomplete sensor payload; missing bands: {', '.join(missing)}")

    for band, value in parsed.items():
        if not math.isfinite(value) or value < 0:
            raise SensorError(f"Band {band} must be a finite, non-negative value")
    return {band: parsed[band] for band in BANDS}


def analyze_bands(values: Mapping[str, float]) -> dict[str, float | str]:
    """Calculate the README-defined red-edge/NIR normalized difference LPQI."""
    red_edge_average = (float(values["S"]) + float(values["T"])) / 2.0
    nir_average = (float(values["U"]) + float(values["V"]) + float(values["W"])) / 3.0
    denominator = nir_average + red_edge_average
    if denominator <= 0:
        raise SensorError("LPQI cannot be calculated because red-edge and NIR values sum to zero")

    lpqi = (nir_average - red_edge_average) / denominator
    quality = "Good" if lpqi >= 0.45 else "Moderate" if lpqi >= 0.25 else "Poor"
    return {
        "red_edge_average": red_edge_average,
        "nir_average": nir_average,
        "lpqi": lpqi,
        "quality": quality,
    }


@dataclass(frozen=True)
class Settings:
    sensor_url: str
    excel_file: Path
    host: str = "127.0.0.1"
    port: int = 5050
    timeout_seconds: float = 10.0
    web_root: Path = Path(__file__).resolve().parent

    @classmethod
    def from_environment(cls) -> "Settings":
        project_dir = Path(__file__).resolve().parent
        sensor_url = os.getenv("MORINGA_SENSOR_URL", "http://10.149.144.149/scan")
        if not sensor_url.rstrip("/").endswith("/scan"):
            sensor_url = sensor_url.rstrip("/") + "/scan"
        return cls(
            sensor_url=sensor_url,
            excel_file=Path(os.getenv(
                "MORINGA_EXCEL_FILE",
                str(project_dir / "outputs" / "moringa_sensor_integration" / "sensor_data.xlsx"),
            )).expanduser(),
            host=os.getenv("MORINGA_HOST", "127.0.0.1"),
            port=int(os.getenv("MORINGA_PORT", "5050")),
            timeout_seconds=float(os.getenv("MORINGA_SENSOR_TIMEOUT", "10")),
            web_root=project_dir,
        )


class SensorClient:
    def __init__(self, url: str, timeout_seconds: float) -> None:
        self.url = url
        self.timeout_seconds = timeout_seconds

    def scan(self) -> dict[str, float]:
        try:
            request = Request(self.url, headers={"Accept": "application/json, text/plain"})
            with urlopen(request, timeout=self.timeout_seconds) as response:
                payload = response.read().decode("utf-8", errors="replace")
        except TimeoutError as exc:
            raise SensorError("The sensor timed out. Check its power, Wi-Fi, and IP address.") from exc
        except HTTPError as exc:
            raise SensorError(f"Sensor returned HTTP {exc.code}.") from exc
        except URLError as exc:
            reason = "timed out" if isinstance(exc.reason, TimeoutError) else str(exc.reason)
            raise SensorError(f"Cannot connect to the sensor at {self.url}: {reason}") from exc
        return parse_sensor_payload(payload)


class ExcelStore:
    """Thread-safe append/read access to the persistent workbook."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self.lock = threading.RLock()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with self.lock:
            self.workbook, self.sheet = self._load_or_create()
            self._ensure_schema()
            self._save()

    def _load_or_create(self):
        if self.path.exists():
            workbook = openpyxl.load_workbook(self.path)
        else:
            workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sensor Readings"
        return workbook, sheet

    def _ensure_schema(self) -> None:
        for column, header in enumerate(HEADERS, start=1):
            self.sheet.cell(1, column, header)

        header_fill = PatternFill("solid", fgColor="006B2C")
        for cell in self.sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.sheet.freeze_panes = "A2"
        self.sheet.auto_filter.ref = f"A1:N{max(1, self.sheet.max_row)}"
        self.sheet.row_dimensions[1].height = 30
        widths = (12, 13, 11, 14, 14, 14, 14, 14, 14, 16, 14, 11, 12, 34)
        for index, width in enumerate(widths, start=1):
            self.sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

        for row in range(2, self.sheet.max_row + 1):
            if all(isinstance(self.sheet.cell(row, col).value, (int, float)) for col in range(4, 10)):
                values = {band: float(self.sheet.cell(row, col).value) for band, col in zip(BANDS, range(4, 10))}
                analysis = analyze_bands(values)
                self.sheet.cell(row, 10, analysis["red_edge_average"])
                self.sheet.cell(row, 11, analysis["nir_average"])
                self.sheet.cell(row, 12, analysis["lpqi"])
                self.sheet.cell(row, 13, analysis["quality"])
                if not self.sheet.cell(row, 14).value:
                    self.sheet.cell(row, 14, "Imported reading")
            for col in range(4, 12):
                self.sheet.cell(row, col).number_format = "0.0000"
            self.sheet.cell(row, 12).number_format = "0.000"

        if not self.sheet.tables:
            table = Table(displayName="SensorReadings", ref=f"A1:N{max(2, self.sheet.max_row)}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            self.sheet.add_table(table)
        else:
            for table in self.sheet.tables.values():
                table.ref = f"A1:N{max(2, self.sheet.max_row)}"

    def _save(self) -> None:
        try:
            self.workbook.save(self.path)
        except PermissionError as exc:
            raise SensorError(
                f"Cannot write {self.path.name}. Close it in Excel and scan again."
            ) from exc

    def _next_sample_number(self) -> int:
        numbers = [
            value for (value,) in self.sheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if isinstance(value, (int, float))
        ]
        return int(max(numbers, default=0)) + 1

    def append(self, values: Mapping[str, float], analysis: Mapping[str, float | str], sensor_url: str) -> dict[str, Any]:
        with self.lock:
            now = datetime.now()
            sample_number = self._next_sample_number()
            self.sheet.append([
                sample_number,
                now.strftime("%Y-%m-%d"),
                now.strftime("%H:%M:%S"),
                *(values[band] for band in BANDS),
                analysis["red_edge_average"],
                analysis["nir_average"],
                analysis["lpqi"],
                analysis["quality"],
                sensor_url,
            ])
            row = self.sheet.max_row
            for col in range(4, 12):
                self.sheet.cell(row, col).number_format = "0.0000"
            self.sheet.cell(row, 12).number_format = "0.000"
            self.sheet.auto_filter.ref = f"A1:N{row}"
            for table in self.sheet.tables.values():
                table.ref = f"A1:N{row}"
            self._save()
            return self._row_to_record(row)

    def _row_to_record(self, row: int) -> dict[str, Any]:
        cells = [self.sheet.cell(row, col).value for col in range(1, 15)]
        return {
            "sample_no": cells[0], "date": str(cells[1] or ""), "time": str(cells[2] or ""),
            "R": cells[3], "S": cells[4], "T": cells[5], "U": cells[6], "V": cells[7], "W": cells[8],
            "red_edge_average": cells[9], "nir_average": cells[10], "lpqi": cells[11],
            "quality": cells[12], "sensor_url": cells[13] or "",
        }

    def history(self, limit: int = 500) -> list[dict[str, Any]]:
        with self.lock:
            start = max(2, self.sheet.max_row - limit + 1)
            records = [self._row_to_record(row) for row in range(start, self.sheet.max_row + 1)]
            return [record for record in reversed(records) if record["sample_no"] is not None]

    def count(self) -> int:
        with self.lock:
            return sum(
                1 for (value,) in self.sheet.iter_rows(min_row=2, max_col=1, values_only=True)
                if value is not None
            )

    def export_bytes(self) -> bytes:
        with self.lock:
            return self.path.read_bytes()

    def clear(self) -> None:
        with self.lock:
            if self.sheet.max_row > 1:
                self.sheet.delete_rows(2, self.sheet.max_row - 1)
            for table in self.sheet.tables.values():
                table.ref = "A1:N2"
            self.sheet.auto_filter.ref = "A1:N1"
            self._save()


class MoringaApplication:
    def __init__(self, settings: Settings, sensor: SensorClient | None = None) -> None:
        self.settings = settings
        self.sensor = sensor or SensorClient(settings.sensor_url, settings.timeout_seconds)
        self.store = ExcelStore(settings.excel_file)
        self.last_scan_ok: bool | None = None
        self.last_scan_error: str | None = None
        self.last_scan_time: str | None = None
        self.scan_lock = threading.Lock()

    def scan(self) -> dict[str, Any]:
        if not self.scan_lock.acquire(blocking=False):
            raise SensorError("A sensor scan is already in progress")
        try:
            values = self.sensor.scan()
            analysis = analyze_bands(values)
            record = self.store.append(values, analysis, self.settings.sensor_url)
            self.last_scan_ok, self.last_scan_error = True, None
            self.last_scan_time = f"{record['date']} {record['time']}"
            return {"ok": True, **record}
        except SensorError as exc:
            self.last_scan_ok, self.last_scan_error = False, str(exc)
            self.last_scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            raise
        finally:
            self.scan_lock.release()

    def status(self) -> dict[str, Any]:
        return {
            "ok": True,
            "sensor_url": self.settings.sensor_url,
            "excel_file": str(self.settings.excel_file),
            "samples": self.store.count(),
            "last_scan_ok": self.last_scan_ok,
            "last_scan_error": self.last_scan_error,
            "last_scan_time": self.last_scan_time,
        }


def make_handler(application: MoringaApplication):
    class APIHandler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(application.settings.web_root), **kwargs)

        def _json(self, status: HTTPStatus, payload: Mapping[str, Any]) -> None:
            body = json.dumps(payload, allow_nan=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def _route(self) -> str:
            return urlparse(self.path).path.rstrip("/") or "/"

        def do_OPTIONS(self) -> None:
            self._json(HTTPStatus.OK, {"ok": True})

        def do_GET(self) -> None:
            route = self._route()
            if route in ("/api/status", "/status"):
                self._json(HTTPStatus.OK, application.status())
            elif route in ("/api/history", "/history"):
                self._json(HTTPStatus.OK, {"ok": True, "rows": application.store.history()})
            elif route in ("/api/scan", "/scan"):
                self._handle_scan()
            elif route == "/api/export":
                try:
                    data = application.store.export_bytes()
                except OSError as exc:
                    self._json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
                    return
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="sensor_data.xlsx"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            else:
                super().do_GET()

        def do_POST(self) -> None:
            if self._route() in ("/api/scan", "/scan"):
                self._handle_scan()
            else:
                self._json(HTTPStatus.NOT_FOUND, {"ok": False, "error": "Not found"})

        def do_DELETE(self) -> None:
            if self._route() in ("/api/history", "/clear"):
                application.store.clear()
                self._json(HTTPStatus.OK, {"ok": True})
            else:
                self._json(HTTPStatus.NOT_FOUND, {"ok": False, "error": "Not found"})

        def _handle_scan(self) -> None:
            try:
                self._json(HTTPStatus.OK, application.scan())
            except SensorError as exc:
                self._json(HTTPStatus.BAD_GATEWAY, {"ok": False, "error": str(exc)})

        def log_message(self, format_string: str, *args: Any) -> None:
            print(f"{self.address_string()} - {format_string % args}")

    return APIHandler


def build_parser() -> argparse.ArgumentParser:
    defaults = Settings.from_environment()
    parser = argparse.ArgumentParser(description="Moringa sensor gateway and Excel logger")
    parser.add_argument("--sensor-url", default=defaults.sensor_url, help="ESP scan URL")
    parser.add_argument("--excel-file", type=Path, default=defaults.excel_file, help="Workbook path")
    parser.add_argument("--host", default=defaults.host)
    parser.add_argument("--port", type=int, default=defaults.port)
    parser.add_argument("--timeout", type=float, default=defaults.timeout_seconds)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    settings = Settings(
        sensor_url=args.sensor_url,
        excel_file=args.excel_file.resolve(),
        host=args.host,
        port=args.port,
        timeout_seconds=args.timeout,
        web_root=Path(__file__).resolve().parent,
    )
    application = MoringaApplication(settings)
    server = ThreadingHTTPServer((settings.host, settings.port), make_handler(application))
    print("=" * 62)
    print("MORINGA NIR SENSOR — LIVE ANALYSIS SERVER")
    print(f"Dashboard : http://localhost:{settings.port}/dashboard.html")
    print(f"Sensor    : {settings.sensor_url}")
    print(f"Workbook  : {settings.excel_file}")
    print("Press Ctrl+C to stop.")
    print("=" * 62)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
